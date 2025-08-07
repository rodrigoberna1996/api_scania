"""
generate_excel_report.py
────────────────────────────────────────────────────────────────────────────
• Construye el reporte Excel de costos de viaje.
• Integra: reasignaciones, viajes vacíos, peajes y datos de Scania
  (km, diésel y AdBlue) SOLO para los viajes exportados.
• Maneja time-outs del API Scania y evita peticiones duplicadas
  mediante un pequeño caché in-memory.
"""

import io
import asyncio
import logging
from asyncio import Semaphore, gather
from typing import List, Any

import httpx
import pandas as pd
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

# ─── Servicios propios ───────────────────────────────────────────────
from app.services.reporting_service.repository import (
    get_filtered_logs,
    get_reassignment_by_title,
)
from app.services.sharepoint_auth.ms_graph import (
    leer_excel_desde_onedrive,
    leer_diesel_desde_onedrive,
    leer_factores_desde_onedrive,
)
from app.services.scania_vehicles.vehicle_map import get_vehicle_map
from app.services.scania_vehicles_status.service import get_vehicle_historical_data

logger = logging.getLogger(__name__)


async def generate_excel_report(session: AsyncSession, mes: int) -> StreamingResponse:
    # ╔════════════════ 1. VIAJES + REASIGNACIONES ══════════════════════╗
    records = await get_filtered_logs(session)
    data = [r.fields for r in records]

    fechas_reasig, horas_reasig = {}, {}
    fechas_desc_real, horas_desc_real = {}, {}
    tracto_reasig: dict[str, Any] = {}
    fecha_desc_prin, hora_desc_prin = {}, {}

    for row in records:
        f = row.fields
        t = f.get("Title")

        # descarga original del viaje
        fecha_desc_prin[t] = pd.to_datetime(f.get("field_16"), dayfirst=True, errors="coerce")
        h_raw = pd.to_datetime(f.get("field_17"), errors="coerce")
        hora_desc_prin[t] = None if pd.isna(h_raw) else h_raw.strftime("%H:%M:%S")

        if f.get("REASIGNACION"):
            reas = await get_reassignment_by_title(session, t)
            if reas:
                ts_reas = pd.to_datetime(
                    reas.fields.get("fecha_reasignacion"), dayfirst=True, errors="coerce"
                )
                if pd.notna(ts_reas):
                    fechas_reasig[t] = ts_reas
                    horas_reasig[t] = ts_reas.time()

                ts_desc_r = pd.to_datetime(
                    reas.fields.get("fecha_descarga_real"), dayfirst=True, errors="coerce"
                )
                if pd.notna(ts_desc_r):
                    fechas_desc_real[t] = ts_desc_r
                    horas_desc_real[t] = ts_desc_r.time()

                tracto_reasig[t] = {
                    "NO_TRACTO":       reas.fields.get("no_tracto"),
                    "PLACAS_TRACTO":   reas.fields.get("placas_tracto"),
                    "NO_REMOLQUE":     reas.fields.get("no_caja"),
                    "PLACAS_REMOLQUE": reas.fields.get("placas_caja"),
                    "NOMBRE_OP":       reas.fields.get("operador"),
                    "ORIGEN":          reas.fields.get("origen"),
                    # DESTINO se conserva
                }

    # ╠════════════════ 2. DATAFRAME BASE ══════════════════════════════╣
    df = (
        pd.DataFrame(data)
        .drop(columns=["@odata.etag"], errors="ignore")
    )

    df["field_16"] = df.apply(
        lambda r: fechas_reasig.get(r["Title"])
        if r["Title"] in fechas_reasig
        else pd.to_datetime(r["field_16"], dayfirst=True, errors="coerce"),
        axis=1,
    )

    df = df.rename(
        columns={
            "field_6": "fecha_carga",
            "field_16": "fecha_descarga",
            "field_1": "No. Económico",
            "field_7": "hora_carga",
            "field_17": "hora_descarga",
            "Title": "Title",
        }
    )

    df["fecha_carga"]    = pd.to_datetime(df["fecha_carga"],    dayfirst=True, errors="coerce")
    df["fecha_descarga"] = pd.to_datetime(df["fecha_descarga"], dayfirst=True, errors="coerce")

    def hora_fmt(row, campo, title, use_reasig=False):
        if use_reasig and title in horas_reasig:
            return horas_reasig[title].strftime("%H:%M:%S")
        h = pd.to_datetime(row[campo], errors="coerce")
        return None if pd.isna(h) else h.strftime("%H:%M:%S")

    df["hora_descarga"] = df.apply(
        lambda r: hora_fmt(r, "hora_descarga", r["Title"], use_reasig=True), axis=1
    )
    df["hora_carga"] = df.apply(
        lambda r: hora_fmt(r, "hora_carga", r["Title"], use_reasig=False), axis=1
    )

    # uniforma ECO
    df["No. Económico"] = df["No. Económico"].apply(
        lambda x: f"ECO {str(x).replace('ECO', '').strip()}"
    )

    # ╠═══════════════ 3. PEAJES ════════════════════════════════════════╣
    peajes_df = await leer_excel_desde_onedrive("Peajes.xlsx", header_row=8)
    peajes_df = peajes_df.iloc[8:]
    peajes_df["Fecha"]         = pd.to_datetime(peajes_df["Fecha"], dayfirst=True, errors="coerce")
    peajes_df["No. Económico"] = peajes_df["No. Económico"].str.strip()
    peajes_df["Costo final"]   = pd.to_numeric(peajes_df["Costo final"], errors="coerce")

    def costo_peajes(r):
        """Devuelve el costo total de peajes para el rango indicado en la fila.

        La función acepta filas en el formato original del DataFrame base
        (columnas 'fecha_carga', 'hora_carga', ...) o en el formato ya
        transformado del DataFrame final ('FECHA_CARGA', 'HORA_CARGA', ...).
        """

        eco = r.get("No. Económico") or r.get("NO_TRACTO")
        fecha_carga = r.get("fecha_carga") or r.get("FECHA_CARGA")
        hora_carga = r.get("hora_carga") or r.get("HORA_CARGA")
        fecha_descarga = r.get("fecha_descarga") or r.get("FECHA_DESCARGA")
        hora_descarga = r.get("hora_descarga") or r.get("HORA_DESCARGA")

        try:
            ini = pd.to_datetime(f"{pd.to_datetime(fecha_carga).date()} {hora_carga}")
            fin = pd.to_datetime(f"{pd.to_datetime(fecha_descarga).date()} {hora_descarga}")
        except Exception:
            return 0

        sel = peajes_df[
            (peajes_df["No. Económico"] == eco)
            & (peajes_df["Fecha"].between(ini, fin))
        ]
        return sel["Costo final"].sum()

    df["PEAJES_VIAPASS"] = (df.apply(costo_peajes, axis=1) / 1.16).round(2)
    df["PEAJES_EFECTIVO"] = (df["PEAJES_EFECTIVO"].fillna(0).astype(float) / 1.16).round(2)
    df["TOTAL_PEAJES"] = (df["PEAJES_VIAPASS"] + df["PEAJES_EFECTIVO"]).round(2)

    # ╠═══════════════ 3-bis. DIESEL ═══════════════════════════════════════╣
    # Lee Diesel.xlsx y prepara un lookup de precios sin IVA
    diesel_df = await leer_diesel_desde_onedrive()

    # Normaliza encabezados y tipos para facilitar las búsquedas por fecha
    diesel_df.columns = diesel_df.columns.str.strip().str.upper()
    if "FECHA" in diesel_df.columns:
        diesel_df["FECHA"] = pd.to_datetime(
            diesel_df["FECHA"], dayfirst=True, errors="coerce"
        )
        diesel_df = diesel_df.sort_values("FECHA").reset_index(drop=True)

    # ── helper de lookup (precio SIN IVA) ────────────────────────────────
    def _precio_diesel_por_fecha(fecha: str | pd.Timestamp) -> float | None:
        if pd.isna(fecha):
            return None
        fecha = pd.to_datetime(fecha).normalize()
        rows = diesel_df[diesel_df["FECHA"] <= fecha]
        if rows.empty:
            return None
        return float(rows.iloc[-1]["PRECIO_DIESEL"])

        # ╠═══════════════ 4. ORDEN Y MAPEOS ════════════════════════════════╣
    df["eco_num"] = pd.to_numeric(
        df["No. Económico"].str.replace("ECO ", "", regex=False), errors="coerce"
    )
    df = (
        df[df["eco_num"].notna()]
        .sort_values(["fecha_carga", "hora_carga"])
        .drop(columns=["eco_num"])
    )

    mapeo = {
        "Title": "TR_NO_VIAJE", "No. Económico": "NO_TRACTO",
        "field_2": "PLACAS_TRACTO", "NO_REMOLQUE": "NO_REMOLQUE",
        "field_3": "PLACAS_REMOLQUE", "field_4": "NOMBRE_OP",
        "ORIGEN_TAB": "ORIGEN", "DESTINO_TAB": "DESTINO",
        "field_8": "CLIENTE", "field_9": "EMPRESA", "CARGA_KILOS": "CARGA_KILOS",
        "field_22": "ADRH_OT", "fecha_carga": "FECHA_CARGA",
        "hora_carga": "HORA_CARGA", "fecha_descarga": "FECHA_DESCARGA",
        "hora_descarga": "HORA_DESCARGA", "REPARTOS1": "REPARTOS",
        "field_19": "MANIOBRAS", "field_20": "ESTADIAS",
        "field_15": "COSTO_VIAJE",
        "COMISION_CLIENTE": "COMISION_CLIENTE",
        "COMISION_OPERADOR": "COMISION_OPERADOR",
        "GASTOS_OPERADOR": "GASTOS_OPERADOR",
        "PEAJES_VIAPASS": "PEAJES_VIAPASS",
    }
    df = df.rename(columns=mapeo)

    cols = [
        "TR_NO_VIAJE","NO_TRACTO","PLACAS_TRACTO","NO_REMOLQUE","PLACAS_REMOLQUE",
        "NOMBRE_OP","ORIGEN","DESTINO","CLIENTE","EMPRESA","CARGA_KILOS","ADRH_OT",
        "FECHA_CARGA","HORA_CARGA","FECHA_DESCARGA","HORA_DESCARGA",
        "REPARTOS","MANIOBRAS","ESTADIAS","FLETE_VACIO","FLETE_FALSO","RECHAZOS",
        "COSTO_VIAJE","KM_RECORRIDOS","CONSUMO_LTS_DIESEL","RENDIMIENTO",
        "PRECIO_DIESEL","COSTO_DIESEL","LTS_ADBLUE_CONSUMIDOS","PRECIO_ADBLUE",
        "COSTO_ADBLUE","COMISION_CLIENTE","COMISION_OPERADOR","GASTOS_OPERADOR",
        "PEAJES_VIAPASS","PEAJES_EFECTIVO","TOTAL_PEAJES","MANTTO_TRACTOS",
        "MANTTO_CAJAS","RASTREO","SEGURO","LLANTAS","ADMINISTRACION",
        "MARKETING","COSTO_TOTAL","UTILIDAD_BRUTA","INGRESO_X_KM","COSTO_X_KM",
        "UTILIDAD_X_KM"
    ]
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    df = df[cols]

    # ╠═══════════════ 5. DUPLICADOS DE REASIGNACIÓN ════════════════════╣
    df["ES_REASIG"] = False
    dup = []
    for _, fila in df.iterrows():
        t = fila["TR_NO_VIAJE"]
        if t in tracto_reasig:
            d = fila.copy()
            info = tracto_reasig[t]
            d["NO_TRACTO"]       = f"ECO {str(info['NO_TRACTO']).strip()}"
            d["PLACAS_TRACTO"]   = info["PLACAS_TRACTO"]
            d["NO_REMOLQUE"]     = info["NO_REMOLQUE"]
            d["PLACAS_REMOLQUE"] = info["PLACAS_REMOLQUE"]
            d["NOMBRE_OP"]       = info["NOMBRE_OP"]
            d["ORIGEN"]          = info["ORIGEN"]
            d["FECHA_CARGA"]     = fechas_desc_real.get(t)
            d["HORA_CARGA"]      = (
                horas_desc_real.get(t).strftime("%H:%M:%S")
                if horas_desc_real.get(t) else None
            )
            d["FECHA_DESCARGA"]  = fecha_desc_prin[t]
            d["HORA_DESCARGA"]   = hora_desc_prin[t]
            d["ES_REASIG"]       = True
            for c in [
                "COSTO_VIAJE",
                "COMISION_CLIENTE",
                "COMISION_OPERADOR",
                "GASTOS_OPERADOR",
                "PEAJES_EFECTIVO",
            ]:
                d[c] = 0
            dup.append(d)
    if dup:
        df = pd.concat([df, pd.DataFrame(dup)], ignore_index=True)

    df["FECHA_CARGA"]    = pd.to_datetime(df["FECHA_CARGA"],    errors="coerce")
    df["FECHA_DESCARGA"] = pd.to_datetime(df["FECHA_DESCARGA"], errors="coerce")

    # ╠═══════════════ 6. VIAJES VACÍOS ═════════════════════════════════╣
    # ── helper para normalizar hora a HH:MM:SS ────────────────────────
    def _hora_to_hms(v: str | None) -> str:
        """Convierte '', None, NaN o 'HH:MM' a 'HH:MM:SS' para to_timedelta."""
        if pd.isna(v) or v in ("", None):
            return "00:00:00"
        v = str(v)
        partes = v.split(":")
        if len(partes) == 1:
            return f"{partes[0]}:00:00"
        if len(partes) == 2:
            return f"{v}:00"
        return v

    df["hora_sort"] = pd.to_timedelta(
        df["HORA_CARGA"].apply(_hora_to_hms)
    )

    viajes = (
        df.sort_values(["NO_TRACTO", "FECHA_CARGA", "hora_sort"])
        .drop(columns=["hora_sort"])
    )

    rows_out: List[pd.Series] = []
    for _, v in viajes.iterrows():
        tracto, fc, hc = v["NO_TRACTO"], v["FECHA_CARGA"], v["HORA_CARGA"]
        prev = viajes[
            (viajes["NO_TRACTO"] == tracto)
            & ((viajes["FECHA_CARGA"] < fc) |
               ((viajes["FECHA_CARGA"] == fc) & (viajes["HORA_CARGA"] < hc)))
        ].sort_values(["FECHA_CARGA", "HORA_CARGA"], ascending=False).head(1)

        vac = pd.Series("", index=cols)
        vac[["KM_RECORRIDOS", "CONSUMO_LTS_DIESEL", "LTS_ADBLUE_CONSUMIDOS"]] = None

        if not prev.empty:
            p = prev.iloc[0]
            vac[["NO_TRACTO","PLACAS_TRACTO","NO_REMOLQUE","PLACAS_REMOLQUE","NOMBRE_OP"]] = \
                p[["NO_TRACTO","PLACAS_TRACTO","NO_REMOLQUE","PLACAS_REMOLQUE","NOMBRE_OP"]]
            vac["ORIGEN"], vac["DESTINO"] = p["DESTINO"], v["ORIGEN"]
            vac["FECHA_CARGA"], vac["HORA_CARGA"] = p["FECHA_DESCARGA"], p["HORA_DESCARGA"]
        else:
            vac[["NO_TRACTO","PLACAS_TRACTO","NO_REMOLQUE","PLACAS_REMOLQUE","NOMBRE_OP"]] = \
                v[["NO_TRACTO","PLACAS_TRACTO","NO_REMOLQUE","PLACAS_REMOLQUE","NOMBRE_OP"]]
            vac["ORIGEN"], vac["DESTINO"] = v["ORIGEN"], v["DESTINO"]

        vac["FECHA_DESCARGA"], vac["HORA_DESCARGA"] = fc, hc
        vac["CLIENTE"] = vac["EMPRESA"] = "VIAJE VACÍO"
        vac["CARGA_KILOS"], vac["TR_NO_VIAJE"] = 0, v["TR_NO_VIAJE"]

        rows_out.extend([vac, v])

    df_final = pd.DataFrame(rows_out).drop(columns=["ES_REASIG"])

    # ╠═══════════════ 6-bis. PEAJES PARA VIAJES VACÍOS ══════════════════════╣
    # Recalcula peajes donde aún no hay valor (los viajes vacíos vienen sin PEAJES_VIAPASS)

    # Asegura que las columnas existen y están como numéricas
    for col in ["PEAJES_VIAPASS", "PEAJES_EFECTIVO"]:
        if col not in df_final.columns:
            df_final[col] = 0
        df_final[col] = pd.to_numeric(df_final[col], errors="coerce")

    # Filtra viajes vacíos (que vienen sin PEAJES_VIAPASS) y recalcula peajes
    mask_vacios = df_final["CLIENTE"] == "VIAJE VACÍO"

    if mask_vacios.any():
        df_final.loc[mask_vacios, "PEAJES_VIAPASS"] = (
            df_final[mask_vacios].apply(costo_peajes, axis=1) / 1.16
        ).round(2)

    # Asegura que PEAJES_EFECTIVO está numérico y rellena vacíos
    df_final["PEAJES_EFECTIVO"] = df_final["PEAJES_EFECTIVO"].fillna(0).astype(float)

    # TOTAL_PEAJES = Viapass + Efectivo
    df_final["TOTAL_PEAJES"] = (
            df_final["PEAJES_VIAPASS"].fillna(0).astype(float) +
            df_final["PEAJES_EFECTIVO"].fillna(0).astype(float)
    ).round(2)

    # ╠═══════════════ 7. FILTRA POR MES ════════════════════════════════╣
    df_final["FECHA_CARGA"] = pd.to_datetime(df_final["FECHA_CARGA"], errors="coerce")
    df_export = df_final[df_final["FECHA_CARGA"].dt.month == mes].copy()
    df_export["ODOMETRO"] = None

    # ╠═══════════════ 8. DATOS SCANIA (km/diesel/adblue) ════════════════╣
    vin_map = await get_vehicle_map()
    sem = Semaphore(3)
    _cache: dict[
        tuple[str, str, str],
        tuple[float | None, float | None, float | None, float | None],
    ] = {}

    def hhmmss(t):
        return "00:00:00" if t is None or pd.isna(t) else str(t).split(" ")[-1][:8]

    async def fetch_scania(idx: int, row: pd.Series):
        eco = str(row["NO_TRACTO"]).replace("ECO", "").strip()
        vin = vin_map.get(eco)
        if (
            not vin
            or pd.isna(row["FECHA_CARGA"])
            or pd.isna(row["FECHA_DESCARGA"])
        ):
            return idx, None, None, None, None

        start = f"{row['FECHA_CARGA'].date()}T{hhmmss(row['HORA_CARGA'])}Z"
        stop  = f"{row['FECHA_DESCARGA'].date()}T{hhmmss(row['HORA_DESCARGA'])}Z"
        key = (vin, start, stop)
        if key in _cache:
            return idx, *_cache[key]

        async with sem:
            try:
                resp = await asyncio.wait_for(
                    get_vehicle_historical_data(vin, start, stop),
                    timeout=20,
                )
                s = resp["summary"]
                km, diesel, adblue, odo = (
                    (
                        s.km_recorridos,
                        s.consumo_lts_diesel,
                        s.lts_adblue_consumidos,
                        s.odometro,
                    )
                    if s
                    else (None, None, None, None)
                )
            except (asyncio.TimeoutError, httpx.TimeoutException, httpx.ReadTimeout):
                logger.warning("Timeout Scania %s  %s–%s", vin, start, stop)
                km = diesel = adblue = odo = None
            except Exception:
                logger.exception("Fallo Scania %s  %s–%s", vin, start, stop)
                km = diesel = adblue = odo = None

        _cache[key] = (km, diesel, adblue, odo)
        return idx, km, diesel, adblue, odo

    tasks = [
        fetch_scania(i, r)
        for i, r in df_export.iterrows()
        if not pd.isna(r["FECHA_CARGA"]) and not pd.isna(r["FECHA_DESCARGA"])
    ]

    for res in await gather(*tasks, return_exceptions=True):
        if isinstance(res, Exception):
            continue
        idx, km, diesel, adblue, odo = res
        if km is not None:
            df_export.at[idx, "KM_RECORRIDOS"]         = round(km, 0)
            df_export.at[idx, "CONSUMO_LTS_DIESEL"]    = round(diesel, 0)
            df_export.at[idx, "LTS_ADBLUE_CONSUMIDOS"] = round(adblue, 2)
            df_export.at[idx, "ODOMETRO"]              = round(odo, 0)

    # ── normaliza cadenas vacías/None a float y rellena 0 ──────────────
    for col in ["KM_RECORRIDOS", "CONSUMO_LTS_DIESEL", "LTS_ADBLUE_CONSUMIDOS"]:
        df_export[col] = pd.to_numeric(df_export[col], errors="coerce").fillna(0).round(2)

    # ╠═══════════════ 8-bis. PRECIO Y COSTO DIÉSEL ════════════════════════╣
    # Calculamos $/L y el costo total de diesel para cada viaje

    df_export["PRECIO_DIESEL"] = df_export["FECHA_CARGA"].apply(_precio_diesel_por_fecha)

    df_export["COSTO_DIESEL"] = (
            pd.to_numeric(df_export["CONSUMO_LTS_DIESEL"], errors="coerce") *
            pd.to_numeric(df_export["PRECIO_DIESEL"], errors="coerce")
    ).round(2)

    # ╠═══════════════ 8-ter. MANTTO TRACTOS ════════════════════════════╣
    factores_df = await leer_factores_desde_onedrive()

    def _factor_por_odometro(odo: float | None) -> float:
        if odo is None or pd.isna(odo):
            return 0.0
        row = factores_df[
            (factores_df["Rango1"] <= odo) & (odo <= factores_df["Rango2"])
        ]
        if row.empty:
            return 0.0
        return float(row.iloc[0]["Factor"])

    df_export["MANTTO_TRACTOS"] = (
        df_export.apply(
            lambda r: _factor_por_odometro(r["ODOMETRO"]) *
            pd.to_numeric(r["KM_RECORRIDOS"], errors="coerce"),
            axis=1,
        )
    ).round(2)
    df_export = df_export.drop(columns=["ODOMETRO"])

    df_export["hora_sort"] = pd.to_timedelta(
        df_export["HORA_CARGA"].apply(_hora_to_hms)
    )

    df_export = (
        df_export
        .sort_values(["FECHA_CARGA", "hora_sort"])
        .drop(columns=["hora_sort"])
        .reset_index(drop=True)
    )

    # ── FORMATO FINAL de fecha y hora ─────────────────────────────────
    for col_fecha in ["FECHA_CARGA", "FECHA_DESCARGA"]:
        df_export[col_fecha] = (
            pd.to_datetime(df_export[col_fecha], errors="coerce")
            .dt.strftime("%Y-%m-%d")
        )

    def _fmt_hora(val):
        if pd.isna(val) or val in ("", None):
            return ""
        if isinstance(val, pd.Timedelta):
            total_min = int(val.total_seconds() // 60)
            return f"{total_min // 60:02d}:{total_min % 60:02d}"
        return str(val)[:5]

    for col_hora in ["HORA_CARGA", "HORA_DESCARGA"]:
        df_export[col_hora] = df_export[col_hora].apply(_fmt_hora)

    # ╠═══════════════ 9. EXPORTA EXCEL ═════════════════════════════════╣
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        df_export.to_excel(xl, index=False, sheet_name="Reporte")
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb["Reporte"]

    green = PatternFill(start_color="99d62b", end_color="99d62b", fill_type="solid")
    gray  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for c in ws[1]:
        c.fill = green
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        if i % 2:
            for c in row:
                c.fill = gray

    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value), default=0) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = w

    final_io = io.BytesIO()
    wb.save(final_io)
    final_io.seek(0)

    nombre_mes = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
        "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ][mes - 1]

    return StreamingResponse(
        final_io,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="Análisis de Costos TR - {nombre_mes}.xlsx"'
        },
    )
